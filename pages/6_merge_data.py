import streamlit as st
import pandas as pd
import openpyxl
import io
import datetime
import re

st.set_page_config(page_title="חיבור מאגרי מידע", page_icon="🔗", layout="wide")

st.title("🔗 חיבור מאגרי מידע (Tabit + YLM)")
st.markdown("""
בעמוד זה תוכל לחבר בין סידור העבודה (Tabit) לבין דוח הנוכחות (YLM) באמצעות טבלת קישור.
המערכת תשתמש בטבלת הקישור כדי לסנכרן בין השמות והעמדות השונים.
""")

# ============================================================
# Helper Functions
# ============================================================

def clean_name_generic(name):
    if not name or pd.isna(name): return ''
    s = str(name).strip()
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'\s*-\s*\d+\s*$', '', s)
    s = re.sub(r'\d{3,}', '', s)
    s = s.replace('*', '').replace('"', '').replace("'", "")
    return ' '.join(s.split()).strip()

def normalize_date(date_val):
    if isinstance(date_val, (datetime.datetime, datetime.date)):
        return date_val.strftime('%d/%m/%Y')
    if not date_val: return ""
    # Try to parse string date
    s = str(date_val).strip()
    match = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', s)
    if match:
        d, m, y = match.groups()
        if len(y) == 2: y = "20" + y
        return f"{int(d):02d}/{int(m):02d}/{y}"
    return s

# ============================================================
# Main UI Logic
# ============================================================

col1, col2, col3 = st.columns(3)

with col1:
    st.subheader("1. סידור עבודה (Tabit)")
    file_tabit = st.file_uploader("העלה את הקובץ המעובד", type=['xlsx'], key="tabit")

with col2:
    st.subheader("2. דוח נוכחות (YLM)")
    file_ylm = st.file_uploader("העלה דוח נוכחות [תאריך, אתר, עובד...]", type=['xlsx', 'xls'], key="ylm")

with col3:
    st.subheader("3. טבלת חיבור (Mappings)")
    file_mapping = st.file_uploader("העלה טבלת קישור שמות ועמדות", type=['xlsx'], key="mapping")

if st.button("🚀 בצע חיבור נתונים", type="primary"):
    if file_tabit and file_ylm and file_mapping:
        try:
            with st.spinner("מעבד נתונים ומבצע חיבור..."):
                # --- A. Load Mapping Table ---
                map_wb = pd.ExcelFile(file_mapping)
                # Sheet 1: Names [שם מסידור העבודה (Tabit), שם בדוח נוכחות (YLM)]
                df_map_names = map_wb.parse(0) 
                # Sheet 2: Positions [עמדת סידור (Tabit), אתרי נוכחות (YLM)]
                df_map_pos = map_wb.parse(1) if len(map_wb.sheet_names) > 1 else pd.DataFrame()

                # Create dictionaries for fast lookup
                name_map = {}
                for _, row in df_map_names.iterrows():
                    k = clean_name_generic(row.iloc[0])
                    v = clean_name_generic(row.iloc[1])
                    if k and v: name_map[k] = v

                pos_map = {}
                if not df_map_pos.empty:
                    for _, row in df_map_pos.iterrows():
                        k = str(row.iloc[0]).strip()
                        v = str(row.iloc[1]).strip()
                        if k and v: pos_map[k] = v

                # --- B. Load Tabit Data ---
                df_tabit = pd.read_excel(file_tabit)
                # Expected: [מקור, שם העובד, תאריך, עמדה, שעת כניסה, שעת יציאה]
                
                # --- C. Load YLM Data ---
                df_ylm = pd.read_excel(file_ylm)
                # Expected: [תאריך, אתר, עובד, כניסה, יציאה, סהכ]

                # 1. Normalize Tabit Names and Positions to YLM language
                df_tabit['שם_נקי_Tabit'] = df_tabit['שם העובד'].apply(clean_name_generic)
                df_tabit['תאריך_נורמלי'] = df_tabit['תאריך'].apply(normalize_date)
                
                # Apply Mapping
                df_tabit['שם_במערכת_YLM'] = df_tabit['שם_נקי_Tabit'].apply(lambda x: name_map.get(x, x))
                df_tabit['אתר_במערכת_YLM'] = df_tabit['עמדה'].apply(lambda x: pos_map.get(str(x).strip(), str(x).strip()))

                # 2. Normalize YLM Data
                df_ylm['שם_נקי_YLM'] = df_ylm['עובד'].apply(clean_name_generic)
                df_ylm['תאריך_נורמלי'] = df_ylm['תאריך'].apply(normalize_date)
                
                # 3. Perform the Merge
                # We merge on Date and standardized YLM Name
                final_merged = pd.merge(
                    df_tabit, 
                    df_ylm, 
                    left_on=['תאריך_נורמלי', 'שם_במערכת_YLM'],
                    right_on=['תאריך_נורמלי', 'שם_נקי_YLM'],
                    how='left', # Keep all Tabit records
                    suffixes=('_סידור', '_נוכחות')
                )

                # Rename columns for clarity before dropping helpers
                final_merged = final_merged.rename(columns={
                    'שם העובד': 'שם בסידור (Tabit)',
                    'עובד': 'שם בנוכחות (YLM)',
                    'עמדה': 'עמדה בסידור (Tabit)',
                    'אתר': 'אתר בנוכחות (YLM)',
                    'שעת כניסה': 'שעת כניסה (Tabit)',
                    'שעת יציאה': 'שעת יציאה (Tabit)',
                    'כניסה': 'כניסה (YLM)',
                    'יציאה': 'יציאה (YLM)'
                })

                # Cleanup: Drop internal helper columns but KEEP the original 'עובד' (now renamed)
                cols_to_drop = ['מקור', 'שם_נקי_Tabit', 'תאריך_נורמלי', 'שם_במערכת_YLM', 'אתר_במערכת_YLM', 'שם_נקי_YLM', 'תאריך_נוכחות', 'ת. כניסה', 'ת. יציאה','סהכ']
                final_merged = final_merged.drop(columns=[c for c in cols_to_drop if c in final_merged.columns])

                # Rename duplicate date column if exists
                if 'תאריך_סידור' in final_merged.columns:
                    final_merged = final_merged.rename(columns={'תאריך_סידור': 'תאריך'})
                elif 'תאריך' not in final_merged.columns and 'תאריך_נוכחות' in final_merged.columns:
                    final_merged = final_merged.rename(columns={'תאריך_נוכחות': 'תאריך'})

                st.success("✅ החיבור הושלם בהצלחה!")
                st.dataframe(final_merged, use_container_width=True)

                # Download
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    final_merged.to_excel(writer, index=False)
                
                st.download_button(
                    label="📥 הורד טבלה מחוברת (Excel)",
                    data=output.getvalue(),
                    file_name=f"merged_data_{datetime.date.today()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

        except Exception as e:
            st.error(f"שגיאה במהלך החיבור: {e}")
            st.info("וודא שהעמודות בקבצים תואמות להגדרות (תאריך, עובד, עמדה וכו')")
    else:
        st.warning("נא להעלות את שלושת הקבצים הנדרשים.")

# ============================================================
# Gap Analysis Section (The "Real Thing")
# ============================================================

st.divider()
st.subheader("🔍 איתור והשלמת פערים (Reconciliation)")
st.markdown("""
העלה כאן את הקובץ המאוחד שנוצר למעלה (או קובץ מחיבור קודם) כדי לאתר עובדים ששכחו להחתים כניסה או יציאה.
המערכת תשלים את השעות החסרות על בסיס הסידור ותפיק דוח צבעוני ומפורט.
""")

merged_file_upload = st.file_uploader("📂 העלה קובץ מאוחד לניתוח פערים", type=['xlsx'], key="reconcile")

if st.button("🚩 נתח פערים והשלם נתונים", type="primary"):
    if merged_file_upload:
        try:
            df_gaps = pd.read_excel(merged_file_upload)
            
            # Check required columns
            required = ['שעת כניסה (Tabit)', 'שעת יציאה (Tabit)', 'כניסה (YLM)', 'יציאה (YLM)']
            missing_cols = [c for c in required if c not in df_gaps.columns]
            
            if missing_cols:
                st.error(f"הקובץ חסר את העמודות הבאות: {', '.join(missing_cols)}")
            else:
                # 1. Identify Gaps
                # A gap is when YLM entry or exit is missing
                mask_entry_missing = df_gaps['כניסה (YLM)'].isna() | (df_gaps['כניסה (YLM)'].astype(str).str.strip() == '')
                mask_exit_missing = df_gaps['יציאה (YLM)'].isna() | (df_gaps['יציאה (YLM)'].astype(str).str.strip() == '')
                
                df_gaps['פער כניסה'] = mask_entry_missing
                df_gaps['פער יציאה'] = mask_exit_missing
                df_gaps['יש פער'] = mask_entry_missing | mask_exit_missing
                
                # 2. Reconcile (Fill Gaps)
                df_gaps['כניסה (YLM) משלים'] = df_gaps['כניסה (YLM)']
                df_gaps.loc[mask_entry_missing, 'כניסה (YLM) משלים'] = df_gaps.loc[mask_entry_missing, 'שעת כניסה (Tabit)']
                
                df_gaps['יציאה (YLM) משלים'] = df_gaps['יציאה (YLM)']
                df_gaps.loc[mask_exit_missing, 'יציאה (YLM) משלים'] = df_gaps.loc[mask_exit_missing, 'שעת יציאה (Tabit)']
                
                # Add a status column for visibility
                def get_status(row):
                    if row['פער כניסה'] and row['פער יציאה']: return "❌ חסר כניסה ויציאה"
                    if row['פער כניסה']: return "⚠️ חסר כניסה"
                    if row['פער יציאה']: return "⚠️ חסר יציאה"
                    return "✅ תקין"
                
                df_gaps['סטטוס סנכרון'] = df_gaps.apply(get_status, axis=1)
                
                # Display results
                st.write(f"### דוח פערים: נמצאו {df_gaps['יש פער'].sum()} שורות עם חוסר סנכרון")
                
                # Filter to show only gaps if requested
                show_only_gaps = st.checkbox("הצג רק שורות עם פערים", value=True)
                display_df = df_gaps[df_gaps['יש פער']] if show_only_gaps else df_gaps
                
                st.dataframe(display_df.drop(columns=['פער כניסה', 'פער יציאה', 'יש פער']), use_container_width=True)
                
                # 3. Create Colorful Excel (Updated with Separators)
                def generate_excel_report(df_source, df_full_with_gaps, sheet_name):
                    output = io.BytesIO()
                    # Group by position and then by date for a clean report
                    df_sorted = df_source.sort_values(by=['עמדה בסידור (Tabit)', 'תאריך'])
                    df_to_save = df_sorted.drop(columns=['פער כניסה', 'פער יציאה', 'יש פער'], errors='ignore')
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        # Create sheet manually to have full control
                        workbook = writer.book
                        worksheet = workbook.create_sheet(sheet_name)
                        if 'Sheet' in workbook.sheetnames: del workbook['Sheet'] # Remove default
                        
                        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                        header_font = Font(color='FFFFFF', bold=True)
                        separator_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid') # Soft Grey separator
                        
                        # Status/YLM Gap colors
                        reported_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') # Green
                        missing_fill = PatternFill(start_color='FABF8F', end_color='FABF8F', fill_type='solid')  # Orange
                        header_gap_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') # Red status
                        
                        # Position palette (Contrasting)
                        POS_COLORS = ['BDD7EE', 'FCE4D6', 'E2EFDA', 'FFF2CC', 'E1E1E1', 'DDEBF7', 'F8CBAD', 'C6E0B4', 'FFE699', 'DAEEF3']
                        unique_positions = df_to_save['עמדה בסידור (Tabit)'].unique()
                        pos_color_map = {pos: PatternFill(start_color=POS_COLORS[i % len(POS_COLORS)], 
                                                        end_color=POS_COLORS[i % len(POS_COLORS)], 
                                                        fill_type='solid') 
                                        for i, pos in enumerate(unique_positions)}
                        
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                        
                        # A. Write Headers
                        for c_idx, col_name in enumerate(df_to_save.columns, start=1):
                            cell = worksheet.cell(row=1, column=c_idx, value=col_name)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                        
                        # Map column indices for styling
                        pos_colored_cols = ['שם בסידור (Tabit)', 'תאריך', 'עמדה בסידור (Tabit)', 'שעת כניסה (Tabit)', 'שעת יציאה (Tabit)']
                        pos_col_indices = [df_to_save.columns.get_loc(c) + 1 for c in pos_colored_cols if c in df_to_save.columns]
                        idx_entry_fixed = df_to_save.columns.get_loc('כניסה (YLM) משלים') + 1
                        idx_exit_fixed = df_to_save.columns.get_loc('יציאה (YLM) משלים') + 1
                        idx_status = df_to_save.columns.get_loc('סטטוס סנכרון') + 1
                        
                        # B. Write Rows with Separators
                        curr_excel_row = 2
                        last_pos = None
                        
                        for _, row_data in df_sorted.iterrows():
                            pos = row_data['עמדה בסידור (Tabit)']
                            
                            # Add separator if position changes
                            if pos != last_pos:
                                worksheet.merge_cells(start_row=curr_excel_row, start_column=1, end_row=curr_excel_row, end_column=len(df_to_save.columns))
                                sep_cell = worksheet.cell(row=curr_excel_row, column=1, value=f"🏢 עמדה: {pos}")
                                sep_cell.font = Font(size=14, bold=True)
                                sep_cell.fill = separator_fill
                                sep_cell.alignment = Alignment(horizontal='center', vertical='center')
                                for c_idx in range(1, len(df_to_save.columns) + 1):
                                    worksheet.cell(row=curr_excel_row, column=c_idx).border = thin_border
                                curr_excel_row += 1
                                last_pos = pos
                            
                            # Write Data Row
                            current_pos_fill = pos_color_map.get(pos)
                            for c_idx, col_name in enumerate(df_to_save.columns, start=1):
                                val = row_data[col_name]
                                cell = worksheet.cell(row=curr_excel_row, column=c_idx, value=val)
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center')
                                
                                # Style Tabit Columns
                                if c_idx in pos_col_indices:
                                    cell.fill = current_pos_fill
                                
                                # Style YLM Fixed Columns
                                if c_idx == idx_entry_fixed:
                                    cell.fill = missing_fill if row_data['פער כניסה'] else reported_fill
                                elif c_idx == idx_exit_fixed:
                                    cell.fill = missing_fill if row_data['פער יציאה'] else reported_fill
                                
                                # Style Status
                                elif c_idx == idx_status:
                                    cell.fill = header_gap_fill if row_data['יש פער'] else reported_fill
                                    
                            curr_excel_row += 1
                        
                        # Auto-adjust width
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if cell.value and len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                                except: pass
                            worksheet.column_dimensions[column_letter].width = max_length + 4
                    return output.getvalue()

                full_report = generate_excel_report(df_gaps, df_gaps, 'דוח פערים והשלמות')
                gaps_only_report = generate_excel_report(df_gaps[df_gaps['יש פער']], df_gaps, 'חוסרים בלבד')

                col_dl1, col_dl2 = st.columns(2)
                with col_dl1:
                    st.download_button(
                        label="📥 הורד דוח מלא (כל העובדים)",
                        data=full_report,
                        file_name=f"Full_Reconciliation_{datetime.date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                with col_dl2:
                    st.download_button(
                        label="📥 הורד דוח חוסרים בלבד (Excel)",
                        data=gaps_only_report,
                        file_name=f"Gaps_Only_Report_{datetime.date.today()}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
        except Exception as e:
            st.error(f"שגיאה בניתוח הפערים: {e}")
    else:
        st.info("אנא העלה את הקובץ המאוחד כדי להתחיל.")
