"""
זיהוי פערי משמרות — Shift Gap Recognition
Standalone Streamlit page: compares Source A (YLM attendance) with Source B (Tabit Shift schedule)
to find guards who forgot to clock in/out.
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from difflib import SequenceMatcher
from io import BytesIO
from datetime import datetime

# ============================================================
# Page Config & CSS
# ============================================================
CUSTOM_CSS = """
<style>
    .main .block-container { direction: rtl; text-align: right; }
    
    /* Step indicator */
    .step-bar { display: flex; gap: 0; margin-bottom: 1.5rem; border-radius: 12px; overflow: hidden; }
    .step-item { flex: 1; padding: 12px 8px; text-align: center; font-size: 14px; font-weight: 600;
                 background: #f0f2f6; color: #888; border-left: 2px solid #fff; transition: all 0.3s;
                 cursor: pointer; }
    .step-item.active { background: linear-gradient(135deg, #667eea, #764ba2); color: #fff; }
    .step-item.done { background: #d4edda; color: #155724; }
    
    /* Match confidence */
    .conf-high { color: #155724; font-weight: bold; }
    .conf-mid { color: #856404; font-weight: bold; }
    .conf-low { color: #721c24; font-weight: bold; }
</style>
"""
st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ============================================================
# Helper Functions
# ============================================================

SHIFT_HOURS = {
    'בוקר': '07:00 – 15:00',
    'צהריים': '15:00 – 23:00',
    'לילה': '23:00 – 07:00',
}

def clean_name_generic(name):
    """
    Robust Hebrew name cleaner.
    1. Removes (נשק), (*), ID numbers.
    2. Trims and joins spaces.
    Precise: Keeps original final letters (ם, ן, etc.) for display.
    """
    if not name or pd.isna(name):
        return ''
    s = str(name).strip()
    
    # Remove text in parentheses
    s = re.sub(r'\(.*?\)', '', s)
    # Remove ID numbers (3+ digits) and trailing dashes with numbers
    s = re.sub(r'\s*-\s*\d+\s*$', '', s)
    s = re.sub(r'\d{3,}', '', s)
    # Remove special chars
    s = s.replace('*', '').replace('"', '').replace("'", "")
    
    return ' '.join(s.split()).strip()

def normalize_final_letters(s):
    """Internal helper to normalize final letters ONLY for comparison."""
    if not s: return ''
    final_map = {'ך': 'כ', 'ם': 'מ', 'ן': 'נ', 'ף': 'פ', 'ץ': 'צ'}
    for f, r in final_map.items():
        s = s.replace(f, r)
    return s

def clean_name_a(name): return clean_name_generic(name)
def clean_name_b(name): return clean_name_generic(name)


def is_time_val(val):
    """Check if value looks like HH:MM."""
    return bool(re.match(r'^\d{1,2}:\d{2}$', str(val).strip())) if val else False


def normalize_date(d):
    """Convert date value to DD/MM/YYYY string."""
    if isinstance(d, datetime):
        return d.strftime('%d/%m/%Y')
    if isinstance(d, pd.Timestamp):
        return d.strftime('%d/%m/%Y')
    s = str(d).strip()
    if re.match(r'^\d{2}/\d{2}/\d{4}$', s):
        return s
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
    return s


# ============================================================
# Parsers
# ============================================================

def parse_source_a(file):
    """Parse Source A (YLM attendance log) into a clean DataFrame."""
    df = pd.read_excel(file)
    cols = df.columns.tolist()
    date_col = next((c for c in cols if 'תאריך' in str(c)), cols[0])
    site_col = next((c for c in cols if 'אתר' in str(c)), cols[1])
    emp_col = next((c for c in cols if 'עובד' in str(c)), cols[2])

    entry_col = None
    exit_col = None
    for c in cols:
        cs = str(c)
        if 'כניסה' in cs and 'ת.' not in cs and 'ת ' not in cs:
            if entry_col is None:  # Take FIRST match only
                entry_col = c
        if 'יציאה' in cs and 'ת.' not in cs and 'ת ' not in cs:
            if exit_col is None:  # Take FIRST match only
                exit_col = c

    df = df[df[emp_col].notna()].copy()
    df['שם_נקי'] = df[emp_col].apply(clean_name_a)
    df = df[df['שם_נקי'] != ''].copy()
    df['תאריך_נקי'] = df[date_col].apply(normalize_date)
    df['אתר'] = df[site_col].astype(str)
    df['יש_כניסה'] = df[entry_col].notna() if entry_col else False
    df['יש_יציאה'] = df[exit_col].notna() if exit_col else False

    def extract_hhmm(val):
        """Extract HH:MM from various datetime/time formats."""
        if pd.isna(val) or val is None:
            return ''
        s = str(val).strip()
        if s in ('', 'nan', 'None', 'NaT'):
            return ''
        m = re.search(r'(\d{1,2}):(\d{2})', s)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return ''

    df['שעת_כניסה'] = df[entry_col].apply(extract_hhmm) if entry_col else ''
    df['שעת_יציאה'] = df[exit_col].apply(extract_hhmm) if exit_col else ''

    return df


def parse_source_b(file):
    """Parse Source B (Tabit Shift work schedule) — complex grid layout.
    
    FIX #2: Improved position detection — captures ALL positions including
    short names and sub-position names (e.g. 'לה"ב 433 בסיסי', 'לה"ב 433').
    A position row is any row in column 0 that is NOT a shift type keyword,
    NOT a time value, NOT entirely numeric, and has actual text content.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    all_records = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Find day header row (cells like "א'\n04/01/2026")
        day_header_idx = None
        dates = {}

        for r_idx, row in enumerate(rows):
            found_in_row = {}
            for c_idx, cell in enumerate(row):
                if cell and isinstance(cell, str) and '\n' in cell:
                    parts = cell.strip().split('\n')
                    date_part = parts[-1].strip()
                    if re.match(r'\d{2}/\d{2}/\d{4}', date_part):
                        found_in_row[c_idx] = date_part
            if len(found_in_row) >= 3:
                day_header_idx = r_idx
                dates = found_in_row
                break

        if day_header_idx is None:
            continue

        day_cols = sorted(dates.keys())
        shift_types = {'בוקר', 'צהריים', 'לילה'}
        current_position = ''
        current_shift_type = ''

        # FIX #2: Pre-scan to detect all position names from column 0
        # A position is any non-empty col0 value that isn't a shift type,
        # time value, or only whitespace/numbers
        all_position_names = set()
        for r_idx in range(day_header_idx + 1, len(rows)):
            row = rows[r_idx]
            if not row:
                continue
            col0 = str(row[0]).strip() if row[0] else ''
            if not col0 or col0 == 'None':
                continue
            if col0 in shift_types:
                continue
            if is_time_val(col0):
                continue
            if re.match(r'^\d+$', col0):
                continue
            # Check if this row has names in day columns (if so, col0 is a sub-position)
            has_names = False
            for dc in day_cols:
                if dc < len(row) and row[dc]:
                    val = str(row[dc]).strip()
                    if val and val != 'None' and not is_time_val(val):
                        has_names = True
                        break
            if has_names:
                # This is a sub-position row (col0 = sub-position label, day cols = names)
                all_position_names.add(col0)
            else:
                # This is a standalone position header row
                all_position_names.add(col0)

        # Now parse with the improved logic
        r_idx = day_header_idx + 1
        while r_idx < len(rows):
            row = rows[r_idx]
            if not row:
                r_idx += 1
                continue

            col0 = str(row[0]).strip() if row[0] else ''

            if col0 in shift_types:
                current_shift_type = col0
                r_idx += 1
                continue

            # Collect names from day columns
            names_in_row = {}
            for dc in day_cols:
                if dc < len(row) and row[dc]:
                    val = str(row[dc]).strip()
                    if val and val != 'None' and not is_time_val(val):
                        names_in_row[dc] = val

            if names_in_row:
                sub_pos = col0 if col0 and col0 not in shift_types else ''

                # Try to get hours from either the next row (Standard) or the previous row (Staggered/Red)
                hours_map = {}  # dc -> (start_t, end_t)

                for dc, raw_name in names_in_row.items():
                    t_start, t_end = '', ''
                    
                    # Target rows to check
                    targets = []
                    if r_idx + 1 < len(rows): targets.append(rows[r_idx + 1])
                    if r_idx - 1 >= 0: targets.append(rows[r_idx - 1])

                    for target_row in targets:
                        if not target_row: continue
                        # Tabit RTL layout mapping: in Python (reading array L to R),
                        # the first time encountered represents ENTRY, the second EXIT.
                        pairs = [(dc, dc + 1), (dc - 1, dc)]
                        for c1, c2 in pairs:
                            if 0 <= c1 < len(target_row) and 0 <= c2 < len(target_row):
                                v1, v2 = str(target_row[c1]).strip(), str(target_row[c2]).strip()
                                if is_time_val(v1) and is_time_val(v2):
                                    t_start, t_end = v1, v2
                                    break
                            elif 0 <= c1 < len(target_row) and is_time_val(str(target_row[c1]).strip()):
                                t_start = str(target_row[c1]).strip()
                                break
                        if t_start: break
                    
                    if t_start:
                        hours_map[dc] = (t_start, t_end)

                for dc, raw_name in names_in_row.items():
                    clean = clean_name_b(raw_name)
                    if clean:
                        start_h, end_h = hours_map.get(dc, ('', ''))
                        # Build shift hours string
                        if start_h and end_h:
                            shift_hours = f"{start_h} – {end_h}"
                        elif start_h:
                            shift_hours = start_h
                        else:
                            shift_hours = SHIFT_HOURS.get(current_shift_type, '')

                        all_records.append({
                            'תאריך': dates.get(dc, ''),
                            'עמדה': current_position,
                            'תת_עמדה': sub_pos,
                            'סוג_משמרת': current_shift_type,
                            'שעות_משמרת': shift_hours,
                            'שם_עובד': clean,
                            'שם_גולמי': raw_name
                        })

                r_idx += 1
                continue

            # Position detection: must not be names, not shift type, not time, not number, 
            # and MUST NOT be a row that looks like an hours row (to avoid overwriting main position)
            elif col0 and col0 != 'None' and col0 not in shift_types and not is_time_val(col0) and not re.match(r'^\d+$', col0):
                # Extra check: is this row actually a times-only row for staggered names?
                has_times_in_row = any(is_time_val(row[dc]) for dc in day_cols if dc < len(row))
                if not has_times_in_row:
                    current_position = col0

            r_idx += 1

    return pd.DataFrame(all_records) if all_records else pd.DataFrame(
        columns=['תאריך', 'עמדה', 'תת_עמדה', 'סוג_משמרת', 'שעות_משמרת', 'שם_עובד', 'שם_גולמי']
    )


# ============================================================
# Fuzzy Matching
# ============================================================

def calc_similarity(s1, s2):
    """
    High-performance fuzzy matching for names.
    Combines: Exact, Substring, Sorted-Token, and Token-Set metrics.
    Normalizes final letters internally for comparison only.
    """
    if not s1 or not s2: return 0.0
    if s1 == s2: return 1.0
    
    # Dynamic normalization for logical comparison
    n1 = normalize_final_letters(s1)
    n2 = normalize_final_letters(s2)
    
    if n1 == n2: return 1.0
    
    # 1. Full Substring Check
    if n1 in n2 or n2 in n1:
        return 0.95

    t1 = n1.split()
    t2 = n2.split()
    
    # 2. Token Set Ratio
    # How many words match regardless of order or extra words?
    set1, set2 = set(t1), set(t2)
    intersection = set1.intersection(set2)
    if intersection:
        # Intersection over the size of the smaller name
        intersect_score = len(intersection) / min(len(set1), len(set2))
        if intersect_score == 1.0: return 0.98  # All words in smaller name exist in larger
    else:
        intersect_score = 0.0

    # 3. Sorted Token SequenceMatcher (Handles order + typos)
    st1 = " ".join(sorted(t1))
    st2 = " ".join(sorted(t2))
    sort_score = SequenceMatcher(None, st1, st2).ratio()
    
    # 4. Standard SequenceMatcher
    orig_score = SequenceMatcher(None, s1, s2).ratio()
    
    # Combine (biased towards Token-Set logic for names)
    return max(intersect_score * 0.9, sort_score, orig_score)


def auto_match(list_a, list_b, threshold=0.6, master_map=None):
    """
    Find best match from list_b for each item in list_a.
    Updated with higher precision logic and tuned threshold.
    """
    if master_map is None: master_map = {}
    matches = []
    # Optimization: Cache cleaned results to avoid repeated work
    for a in list_a:
        if a in master_map and master_map[a] in list_b:
            matches.append({
                'source_a': a,
                'source_b': master_map[a],
                'confidence': 100.0
            })
            continue

        best_b = None
        best_score = 0.0
        for b in list_b:
            score = calc_similarity(a, b)
            if score > best_score:
                best_score = score
                best_b = b
                if score == 1.0: break # Early exit for perfect match
                
        matches.append({
            'source_a': a,
            'source_b': best_b if best_score >= threshold else None,
            'confidence': round(best_score * 100, 1)
        })
    return matches


def auto_match_positions_multi(list_b, list_a, threshold=0.35, master_pos_map=None):
    """Auto-match positions: defaults to the SINGLE best Source A match per Source B,
    but the UI allows selecting multiple via multiselect.
    Returns list of dicts: {source_b, matched_a: [best_match], confidence}
    """
    if master_pos_map is None: master_pos_map = {}
    matches = []
    for b in list_b:
        if b in master_pos_map:
            valid_a = [x for x in master_pos_map[b] if x in list_a]
            if valid_a:
                matches.append({
                    'source_b': b,
                    'matched_a': valid_a,
                    'confidence': 100.0
                })
                continue

        best_a = None
        best_score = 0.0
        for a in list_a:
            score = calc_similarity(b, a)
            if score > best_score:
                best_score = score
                best_a = a
        matches.append({
            'source_b': b,
            'matched_a': [best_a] if best_a and best_score >= threshold else [],
            'confidence': round(best_score * 100, 1)
        })
    return matches


# ============================================================
# Gap Detection Engine
# ============================================================

def _time_to_minutes(t):
    """Convert HH:MM string to minutes since midnight."""
    if not t:
        return -1
    m = re.match(r'^(\d{1,2}):(\d{2})', str(t).strip())
    if not m:
        return -1
    return int(m.group(1)) * 60 + int(m.group(2))

def _parse_absolute_dt(d_str, t_str):
    """Parse a date string DD/MM/YYYY and time HH:MM into a datetime object."""
    try:
        dt = datetime.strptime(d_str, '%d/%m/%Y')
        if t_str:
            m = re.match(r'^(\d{1,2}):(\d{2})', str(t_str).strip())
            if m:
                return dt.replace(hour=int(m.group(1)), minute=int(m.group(2)))
        return dt
    except:
        return None

def detect_gaps(df_a, df_b, name_map, pos_map_multi, deleted_positions=None):
    """
    For each scheduled shift in Source B, check Source A for clock-in/out.
    Uses absolute datetime math so night-shift punches on the next calendar day
    are correctly found and matched to the previous day's shift.
    """
    results = []
    if deleted_positions is None:
        deleted_positions = set()

    # Build lookup: clean_name_a -> list of (punch_datetime, orig_time_str)
    a_person_punches = {}
    for _, row in df_a.iterrows():
        name = row['שם_נקי']
        d_str = row['תאריך_נקי']
        
        if name not in a_person_punches:
            a_person_punches[name] = []
            
        entry_t = str(row.get('שעת_כניסה', '')).strip()
        exit_t = str(row.get('שעת_יציאה', '')).strip()
        
        for t_str in (entry_t, exit_t):
            if t_str and t_str not in ('', 'nan', 'None'):
                punch_dt = _parse_absolute_dt(d_str, t_str)
                if punch_dt:
                    a_person_punches[name].append((punch_dt, t_str))

    for _, row_b in df_b.iterrows():
        b_name = row_b['שם_עובד']
        b_date = row_b['תאריך']
        b_position = row_b.get('עמדה', '')
        b_sub_position = row_b.get('תת_עמדה', '')
        b_shift = row_b.get('סוג_משמרת', '')
        b_hours = row_b.get('שעות_משמרת', SHIFT_HOURS.get(b_shift, ''))

        # Parse scheduled hours
        sch_start, sch_end = '', ''
        if '–' in b_hours:
            parts = b_hours.split('–')
            sch_start = parts[0].strip()
            sch_end = parts[1].strip()
        elif b_hours:
            sch_start = b_hours.split()[0] if b_hours.split() else ''

        # Determine the display position
        display_position = b_sub_position if b_sub_position else b_position

        # Skip only if the DISPLAY position was explicitly deleted by the user
        if display_position in deleted_positions:
            continue

        a_name = name_map.get(b_name)
        best_entry, best_exit = '', ''
        has_entry, has_exit = False, False

        if a_name is None:
            status = '❌ לא החתים כלל'
        else:
            # Construct absolute shift datetimes
            b_dt = _parse_absolute_dt(b_date, None)
            valid_punches = []
            
            if b_dt and sch_start and sch_end:
                sch_start_dt = _parse_absolute_dt(b_date, sch_start)
                sch_end_dt = _parse_absolute_dt(b_date, sch_end)
                
                # If end time is before start time (e.g. 19:00 - 07:00), shift crossed midnight
                from datetime import timedelta
                if sch_end_dt and sch_start_dt and sch_end_dt <= sch_start_dt:
                    sch_end_dt += timedelta(days=1)
                
                if sch_start_dt and sch_end_dt:
                    all_punches = a_person_punches.get(a_name, [])
                    
                    for punch_dt, t_str in all_punches:
                        # Because YLM might use 'Logical Shift Date' instead of calendar date,
                        # a morning exit punch (e.g. 07:11 on Day 2) might be recorded under Day 1.
                        # We evaluate the punch as-is, +1 day, and -1 day to see if it physically maps
                        # to our shift start/end boundaries.
                        
                        best_dist_start = float('inf')
                        best_dist_end = float('inf')
                        best_match_dt = None
                        
                        for days_offset in (0, 1, -1):
                            test_dt = punch_dt + timedelta(days=days_offset)
                            
                            dist_to_start = abs((test_dt - sch_start_dt).total_seconds())
                            dist_to_end = abs((test_dt - sch_end_dt).total_seconds())
                            
                            # If it's within 6 hours (21600 seconds) of either start or end, it's a potential match
                            if dist_to_start <= 21600 or dist_to_end <= 21600:
                                if min(dist_to_start, dist_to_end) < min(best_dist_start, best_dist_end):
                                    best_dist_start = dist_to_start
                                    best_dist_end = dist_to_end
                                    best_match_dt = test_dt
                                    
                        if best_match_dt:
                            valid_punches.append((best_match_dt, t_str, best_dist_start, best_dist_end))

            # Deduplicate punches to avoid same time string twice
            unique_punches = {}
            for p_dt, t_str, d_start, d_end in valid_punches:
                if t_str not in unique_punches:
                    unique_punches[t_str] = (p_dt, t_str, d_start, d_end)
            
            valid_punches = list(unique_punches.values())

            if valid_punches:
                if len(valid_punches) == 1:
                    # One punch: closer to start or end?
                    _, t_str, dist_start, dist_end = valid_punches[0]
                    if dist_start <= dist_end:
                        has_entry = True
                        best_entry = t_str
                    else:
                        has_exit = True
                        best_exit = t_str
                
                elif len(valid_punches) >= 2:
                    # Find closest to start
                    entry_candidate = min(valid_punches, key=lambda x: x[2])
                    # Find closest to end
                    exit_candidate = min(valid_punches, key=lambda x: x[3])
                    
                    if entry_candidate[1] == exit_candidate[1]: # Same punch won both
                        if entry_candidate[2] <= exit_candidate[3]:
                            has_entry = True
                            best_entry = entry_candidate[1]
                            rem = [p for p in valid_punches if p[1] != entry_candidate[1]]
                            if rem:
                                has_exit = True
                                best_exit = min(rem, key=lambda x: x[3])[1]
                        else:
                            has_exit = True
                            best_exit = exit_candidate[1]
                            rem = [p for p in valid_punches if p[1] != exit_candidate[1]]
                            if rem:
                                has_entry = True
                                best_entry = min(rem, key=lambda x: x[2])[1]
                    else:
                        has_entry = True
                        best_entry = entry_candidate[1]
                        has_exit = True
                        best_exit = exit_candidate[1]

                if has_entry and has_exit:
                    status = '✅ תקין'
                elif has_entry and not has_exit:
                    status = '⚠️ חסר יציאה'
                elif not has_entry and has_exit:
                    status = '⚠️ חסר כניסה'
                else:
                    status = '❌ חסר כניסה ויציאה'
            else:
                status = '❌ לא החתים כלל'

        # Prefer actual time, fallback to scheduled
        final_entry = best_entry if has_entry else (sch_start if sch_start else 'X')
        final_exit = best_exit if has_exit else (sch_end if sch_end else 'X')

        results.append({
            'עמדה ראשית': b_position if b_position else display_position,
            'עמדה': display_position,
            'תאריך': b_date,
            'שם עובד': b_name,
            'משמרת': b_shift,
            'שעות (סידור)': b_hours,
            'כניסה': final_entry,
            'יציאה': final_exit,
            'סטטוס': status
        })

    df_result = pd.DataFrame(results)
    # FIX #9: Proper date sorting (DD/MM/YYYY needs parsing for numeric sort)
    if not df_result.empty:
        # Create temp column for sorting
        df_result['_sort_date'] = pd.to_datetime(df_result['תאריך'], dayfirst=True, errors='coerce')
        df_result = df_result.sort_values(['עמדה ראשית', 'עמדה', '_sort_date', 'משמרת']).reset_index(drop=True)
        df_result = df_result.drop(columns=['_sort_date'])
    return df_result


# ============================================================
# Natural Language Gap Report
# ============================================================

def generate_natural_language_report(gap_df):
    """Generate a human-readable text report of gaps, grouped by position."""
    if gap_df is None or gap_df.empty:
        return "לא נמצאו פערים."

    problems = gap_df[~gap_df['סטטוס'].str.contains('תקין')].copy()
    if problems.empty:
        return "🎉 לא נמצאו פערים — כל המאבטחים החתימו כניסה ויציאה כנדרש!"

    lines = []
    # Group by main position
    group_col = 'עמדה ראשית' if 'עמדה ראשית' in problems.columns else 'עמדה'
    grouped = problems.groupby(group_col)

    for position, group in grouped:
        pos_label = position if position else 'עמדה לא ידועה'
        lines.append(f"\n🏢 **{pos_label}**\n")

        for idx, (_, row) in enumerate(group.iterrows(), 1):
            date = row['תאריך']
            name = row['שם עובד']
            shift = row['משמרת']
            hours = row['שעות (סידור)']
            status = row['סטטוס']

            if '⚠️ חסר יציאה' in status:
                lines.append(
                    f"{idx}. ביום **{date}**, **{name}** עשה כניסה למשמרת {shift} "
                    f"({hours}) ב{pos_label} — אבל **שכח לעשות יציאה**."
                )
            elif '⚠️ חסר כניסה' in status:
                lines.append(
                    f"{idx}. ביום **{date}**, **{name}** עשה יציאה ממשמרת {shift} "
                    f"({hours}) ב{pos_label} — אבל **שכח לעשות כניסה**."
                )
            elif 'לא החתים כלל' in status:
                lines.append(
                    f"{idx}. ביום **{date}**, **{name}** שכח להחתים משמרת {shift} "
                    f"ב{pos_label} בין השעות {hours} — **לא עשה כניסה ולא יציאה**."
                )
            elif 'חסר כניסה ויציאה' in status:
                lines.append(
                    f"{idx}. ביום **{date}**, **{name}** שכח להחתים משמרת {shift} "
                    f"ב{pos_label} בין השעות {hours} — **לא עשה כניסה ולא יציאה**."
                )

    return '\n'.join(lines)


# ============================================================
# Rich Excel Export (FIX #4: Big position header separators)
# ============================================================

def generate_rich_excel(gap_df, natural_text):
    """Generate a beautifully styled Excel workbook for gap results.
    FIX #4: Insert a big, prominent position header row between positions.
    """
    buffer = BytesIO()
    wb = openpyxl.Workbook()

    # --- Colors and Styles ---
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    ok_font = Font(name='Arial', color='006100', size=11)
    partial_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    partial_font = Font(name='Arial', color='9C6500', size=11)
    missing_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    missing_font = Font(name='Arial', color='9C0006', size=11)
    normal_font = Font(name='Arial', size=11)
    bold_font = Font(name='Arial', bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD'),
    )
    # FIX #4: Big position separator styles
    pos_header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    pos_header_font = Font(name='Arial', bold=True, size=16, color='FFFFFF')
    pos_header_align = Alignment(horizontal='center', vertical='center')
    pos_header_border = Border(
        left=Side(style='medium', color='1F4E79'),
        right=Side(style='medium', color='1F4E79'),
        top=Side(style='medium', color='1F4E79'),
        bottom=Side(style='medium', color='1F4E79'),
    )
    # Unmatched employee style (blue/info)
    unmatched_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    unmatched_font = Font(name='Arial', color='1F4E79', size=11)

    # Determine columns for display (exclude 'עמדה ראשית' internal column)
    display_columns = ['עמדה', 'תאריך', 'שם עובד', 'משמרת', 'שעות (סידור)', 'כניסה', 'יציאה', 'סטטוס']
    col_widths = [35, 14, 22, 12, 18, 10, 10, 22]
    num_cols = len(display_columns)

    # ============== Sheet 1: All Results ==============
    ws = wb.active
    ws.title = 'דוח פערים'
    ws.sheet_view.rightToLeft = True

    # Set column widths
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # FIX #4: Write data with big position headers
    if not gap_df.empty:
        group_col = 'עמדה ראשית' if 'עמדה ראשית' in gap_df.columns else 'עמדה'
        current_main_pos = None
        row_num = 1

        for _, data_row in gap_df.iterrows():
            main_pos = data_row.get(group_col, '')

            # Insert big position header when main position changes
            if main_pos != current_main_pos:
                current_main_pos = main_pos
                if row_num > 1:
                    row_num += 1  # blank row before header

                # FIX #4: Merge cells for the big position header
                pos_label = main_pos if main_pos else 'עמדה לא ידועה'
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
                header_cell = ws.cell(row=row_num, column=1, value=f"📋  {pos_label}")
                header_cell.fill = pos_header_fill
                header_cell.font = pos_header_font
                header_cell.alignment = pos_header_align
                header_cell.border = pos_header_border
                ws.row_dimensions[row_num].height = 35
                # Apply border to all merged cells
                for c in range(2, num_cols + 1):
                    cell = ws.cell(row=row_num, column=c)
                    cell.fill = pos_header_fill
                    cell.border = pos_header_border
                row_num += 1

                # Column headers after each position header
                for c_idx, col_name in enumerate(display_columns, 1):
                    cell = ws.cell(row=row_num, column=c_idx, value=col_name)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                row_num += 1

            # Data row
            for c_idx, col_name in enumerate(display_columns, 1):
                val = data_row.get(col_name, '')
                cell = ws.cell(row=row_num, column=c_idx, value=val)
                cell.font = normal_font
                cell.alignment = center_align if c_idx != 1 else right_align
                cell.border = thin_border

                # Color code status
                status = str(data_row.get('סטטוס', ''))
                if col_name == 'סטטוס':
                    if '✅' in status:
                        cell.fill = ok_fill
                        cell.font = ok_font
                    elif '⚠️' in status:
                        cell.fill = partial_fill
                        cell.font = partial_font
                    elif '❌' in status:
                        cell.fill = missing_fill
                        cell.font = missing_font
                    elif 'לא שודך' in status:
                        cell.fill = unmatched_fill
                        cell.font = unmatched_font

                # Color code entry/exit relative to status
                status = str(data_row.get('סטטוס', ''))
                if col_name in ('כניסה', 'יציאה'):
                    if '✅' in status:
                        cell.fill = ok_fill
                        cell.font = ok_font
                    elif '⚠️' in status:
                        # If the specific cell is 'X' or matches a missing state in status
                        is_missing = (val == 'X')
                        if (col_name == 'כניסה' and 'חסר כניסה' in status) or \
                           (col_name == 'יציאה' and 'חסר יציאה' in status) or is_missing:
                            cell.fill = missing_fill
                            cell.font = missing_font
                        else:
                            cell.fill = ok_fill
                            cell.font = ok_font
                    elif '❌' in status or 'לא שודך' in status:
                        cell.fill = missing_fill
                        cell.font = missing_font

            row_num += 1

    # ============== Sheet 2: Problems Only ==============
    problems = gap_df[~gap_df['סטטוס'].str.contains('תקין')] if not gap_df.empty else pd.DataFrame()
    if not problems.empty:
        ws2 = wb.create_sheet('בעיות בלבד')
        ws2.sheet_view.rightToLeft = True
        for i, w in enumerate(col_widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        group_col = 'עמדה ראשית' if 'עמדה ראשית' in problems.columns else 'עמדה'
        current_main_pos = None
        row_num = 1

        for _, data_row in problems.iterrows():
            main_pos = data_row.get(group_col, '')

            if main_pos != current_main_pos:
                current_main_pos = main_pos
                if row_num > 1:
                    row_num += 1

                pos_label = main_pos if main_pos else 'עמדה לא ידועה'
                ws2.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
                hcell = ws2.cell(row=row_num, column=1, value=f"📋  {pos_label}")
                hcell.fill = pos_header_fill
                hcell.font = pos_header_font
                hcell.alignment = pos_header_align
                hcell.border = pos_header_border
                ws2.row_dimensions[row_num].height = 35
                for c in range(2, num_cols + 1):
                    cell = ws2.cell(row=row_num, column=c)
                    cell.fill = pos_header_fill
                    cell.border = pos_header_border
                row_num += 1

                for c_idx, col_name in enumerate(display_columns, 1):
                    cell = ws2.cell(row=row_num, column=c_idx, value=col_name)
                    cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
                    cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
                    cell.alignment = center_align
                    cell.border = thin_border
                row_num += 1

            for c_idx, col_name in enumerate(display_columns, 1):
                val = data_row.get(col_name, '')
                cell = ws2.cell(row=row_num, column=c_idx, value=val)
                cell.font = normal_font
                cell.alignment = center_align if c_idx != 1 else right_align
                cell.border = thin_border
                status = str(data_row.get('סטטוס', ''))
                if col_name == 'סטטוס':
                    if '⚠️' in status:
                        cell.fill = partial_fill
                        cell.font = partial_font
                    elif '❌' in status:
                        cell.fill = missing_fill
                        cell.font = missing_font
                # Color code entry/exit relative to status (Problems only sheet)
                if col_name in ('כניסה', 'יציאה'):
                    is_missing = (val == 'X')
                    if (col_name == 'כניסה' and 'חסר כניסה' in status) or \
                       (col_name == 'יציאה' and 'חסר יציאה' in status) or is_missing or 'לא החתים' in status:
                        cell.fill = missing_fill
                        cell.font = missing_font
                    else:
                        cell.fill = ok_fill
                        cell.font = ok_font

            row_num += 1

    # ============== Sheet 3: Natural Language Report ==============
    ws3 = wb.create_sheet('דוח טקסטואלי')
    ws3.sheet_view.rightToLeft = True
    ws3.column_dimensions['A'].width = 120

    ws3.cell(row=1, column=1, value='דוח פערי החתמות — שפה טבעית').font = Font(
        name='Arial', bold=True, size=16, color='1F4E79')

    pos_fill_style = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    # Strip markdown bold markers for Excel
    clean_text = natural_text.replace('**', '')
    for r_idx, line in enumerate(clean_text.split('\n'), 3):
        cell = ws3.cell(row=r_idx, column=1, value=line)
        if line.startswith('🏢'):
            cell.font = Font(name='Arial', bold=True, size=13, color='1F4E79')
            cell.fill = pos_fill_style
        else:
            cell.font = Font(name='Arial', size=11)
        cell.alignment = Alignment(horizontal='right', wrap_text=True)

    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================
# Render Step Indicator
# ============================================================

def render_steps(current):
    labels = ['העלאת קבצים', 'שידוך שמות', 'שידוך עמדות', 'שידוך תאריכים', 'זיהוי פערים']
    icons = ['📁', '👤', '🏢', '📅', '🔍']
    html_parts = []
    for i, (label, icon) in enumerate(zip(labels, icons), 1):
        if i < current:
            cls = 'done'
        elif i == current:
            cls = 'active'
        else:
            cls = ''
        html_parts.append(f'<div class="step-item {cls}">{icon} שלב {i}: {label}</div>')
    st.markdown(f'<div class="step-bar">{"".join(html_parts)}</div>', unsafe_allow_html=True)


# ============================================================
# Session State Initialization
# ============================================================

for key, default in [
    ('master_name_map', {}),
    ('master_pos_map', {}),
    ('gap_step', 1),
    ('df_a', None),
    ('df_b', None),
    ('names_a', []),
    ('names_b', []),
    ('name_matches', []),
    ('name_map_final', {}),
    ('positions_a', []),
    ('positions_b', []),
    ('pos_matches_multi', []),      # FIX #3: multi-match structure
    ('pos_map_multi', {}),          # FIX #3: {pos_b: [pos_a_1, pos_a_2, ...]}
    ('dates_a', []),
    ('dates_b', []),
    ('gap_results', None),
    ('deleted_guards', set()),
    ('deleted_positions', set()),    # Track deleted positions
    # FIX #1: Persist user selections across steps
    ('saved_name_selections', {}),   # {index: selected_value}
    ('saved_pos_selections', {}),    # {index: [selected_values]}
    ('step_completed', set()),       # Track which steps have been completed
]:
    if key not in st.session_state:
        st.session_state[key] = default


# ============================================================
# Main UI
# ============================================================

st.title("🔍 זיהוי פערי משמרות")
st.caption("השוואת שעון נוכחות (יל\"מ) מול סידור עבודה (Tabit Shift) — לזיהוי החתמות חסרות")
st.markdown("---")

current_step = st.session_state['gap_step']
render_steps(current_step)

# ============================================================
# STEP 1: Upload Files
# ============================================================
if current_step == 1:
    st.header("📁 שלב 1: העלאת קבצים")
    st.info("העלה את שני הקבצים: דוח נוכחות מיל\"מ (מאגר A) וסידור עבודה מ-Tabit Shift (מאגר B)")

    st.markdown("#### 🔗 קובץ מיפוי קבוע (אופציונלי)")
    st.info("אם הורדת קובץ מיפוי מ'מחולל המיפויים', העלה אותו כאן כדי לדלג על התאמה ידנית!")
    master_map_file = st.file_uploader("העלה Master_Mapping.xlsx (לא חובה)", type=['xlsx', 'xls'], key='upload_master')
    
    st.markdown("---")
    col_a, col_b = st.columns(2)
    with col_a:
        st.markdown("#### מאגר A — דוח נוכחות (יל\"מ)")
        file_a = st.file_uploader("העלה קובץ Excel", type=['xlsx', 'xls'], key='upload_a')
    with col_b:
        st.markdown("#### מאגר B — סידור עבודה (Tabit Shift)")
        files_b = st.file_uploader("העלה קובץ Excel (ניתן להעלות מספר קבצים)", type=['xlsx', 'xls'], key='upload_b', accept_multiple_files=True)

    if file_a and files_b:
        with st.spinner("מנתח קבצים..."):
            try:
                if master_map_file:
                    try:
                        df_emp_map = pd.read_excel(master_map_file, sheet_name='עובדים')
                        st.session_state['master_name_map'] = dict(zip(
                            df_emp_map['שם מסידור העבודה (Tabit)'], 
                            df_emp_map['שם בדוח נוכחות (YLM)']
                        ))
                    except: pass
                    try:
                        df_pos_map = pd.read_excel(master_map_file, sheet_name='עמדות')
                        pos_dict = {}
                        for _, row in df_pos_map.iterrows():
                            t = row['עמדת סידור (Tabit)']
                            y = row['אתרי נוכחות (YLM)']
                            if not pd.isna(y):
                                pos_dict[t] = [x.strip() for x in str(y).split(',') if x.strip()]
                        st.session_state['master_pos_map'] = pos_dict
                    except: pass

                # Parse Source A (Single file)
                df_a = parse_source_a(file_a)
                
                # Parse and Merge Source B (Multiple files)
                all_b_dfs = []
                for fb in files_b:
                    temp_df_b = parse_source_b(fb)
                    if temp_df_b is not None and not temp_df_b.empty:
                        all_b_dfs.append(temp_df_b)
                
                if not all_b_dfs:
                    st.error("לא ניתן היה לחלץ נתונים מקבצי מאגר B.")
                    st.stop()
                    
                df_b = pd.concat(all_b_dfs, ignore_index=True)
                
                # Deduplicate df_b just in case overlapping dates/names exist across files
                df_b = df_b.drop_duplicates(subset=['תאריך', 'שם_עובד', 'עמדה', 'תת_עמדה', 'סוג_משמרת', 'שעות_משמרת']).reset_index(drop=True)

                st.session_state['df_a'] = df_a
                st.session_state['df_b'] = df_b

                names_a = sorted(df_a['שם_נקי'].unique().tolist())
                names_b = sorted(df_b['שם_עובד'].unique().tolist())
                st.session_state['names_a'] = names_a
                st.session_state['names_b'] = names_b

                positions_a = sorted(df_a['אתר'].unique().tolist())
                # FIX #2: Collect ALL unique positions AND sub-positions from Source B
                all_b_positions = set()
                for p in df_b['עמדה'].unique().tolist():
                    if p and str(p).strip():
                        all_b_positions.add(str(p).strip())
                for p in df_b['תת_עמדה'].unique().tolist():
                    if p and str(p).strip():
                        all_b_positions.add(str(p).strip())
                positions_b = sorted(all_b_positions)

                st.session_state['positions_a'] = positions_a
                st.session_state['positions_b'] = positions_b

                dates_a = sorted(df_a['תאריך_נקי'].unique().tolist())
                dates_b = sorted(df_b['תאריך'].unique().tolist())
                st.session_state['dates_a'] = dates_a
                st.session_state['dates_b'] = dates_b

                st.success(f"✅ מאגר A: {len(df_a)} רשומות, {len(names_a)} עובדים")
                st.success(f"✅ מאגר B: {len(df_b)} רשומות, {len(names_b)} עובדים, {len(positions_b)} עמדות")
            except Exception as e:
                st.error(f"שגיאה בניתוח הקבצים: {e}")
                st.exception(e)

        with st.expander("👁️ תצוגה מקדימה — מאגר A (נוכחות)"):
            st.dataframe(df_a.head(20), use_container_width=True)
        with st.expander("👁️ תצוגה מקדימה — מאגר B (סידור עבודה)"):
            st.dataframe(df_b.head(20), use_container_width=True)
        with st.expander("📋 עמדות שזוהו במאגר B"):
            for p in positions_b:
                st.write(f"• {p}")

        if st.button("המשך לשלב הבא ←", type="primary", use_container_width=True):
            # Only auto-match if we haven't done it before
            if not st.session_state['name_matches']:
                matches = auto_match(names_b, names_a, master_map=st.session_state.get('master_name_map', {}))
                st.session_state['name_matches'] = matches
            st.session_state['step_completed'].add(1)
            st.session_state['gap_step'] = 2
            st.rerun()

# ============================================================
# STEP 2: Name Matching (FIX #1: Persistent state)
# ============================================================
elif current_step == 2:
    st.header("👤 שלב 2: שידוך שמות עובדים")
    st.info("שדך שמות עובדים בין המאגרים. מי שנשאר 'לא שודך' עדיין יופיע בדוח עם השם מסידור העבודה (ללא נתוני החתמה). ניתן גם למחוק מאבטחים מהרשימה.")

    names_a = st.session_state['names_a']
    names_b = st.session_state['names_b']
    matches = st.session_state['name_matches']
    deleted = st.session_state['deleted_guards']

    if not matches:
        matches = auto_match(names_b, names_a, master_map=st.session_state.get('master_name_map', {}))
        st.session_state['name_matches'] = matches

    # Filter out deleted guards
    active_matches = [m for m in matches if m['source_a'] not in deleted]

    options_a = ['— לא שודך —'] + names_a
    name_map = {}

    saved_selections = st.session_state['saved_name_selections']

    for i, m in enumerate(active_matches):
        guard_name = m['source_a']  # Use name as stable key (not index!)
        col1, col2, col3, col4 = st.columns([3, 4, 1.5, 1])
        with col1:
            st.markdown(f"**{guard_name}**")
            st.caption("מאגר B (סידור)")
        with col2:
            # Use saved selection keyed by NAME (stable across deletions)
            if guard_name in saved_selections and saved_selections[guard_name] in options_a:
                default_idx = options_a.index(saved_selections[guard_name])
            elif m['source_b'] and m['source_b'] in options_a:
                default_idx = options_a.index(m['source_b'])
            else:
                default_idx = 0
            selected = st.selectbox(
                f"שידוך ל-{guard_name}",
                options_a,
                index=default_idx,
                key=f"name_match_{guard_name}",
                label_visibility="collapsed"
            )
            # Save selection keyed by name
            saved_selections[guard_name] = selected
            if selected != '— לא שודך —':
                name_map[guard_name] = selected
        with col3:
            conf = m['confidence']
            cls = 'conf-high' if conf >= 80 else ('conf-mid' if conf >= 50 else 'conf-low')
            st.markdown(f'<span class="{cls}">{conf}%</span>', unsafe_allow_html=True)
        with col4:
            if st.button("🗑️", key=f"del_guard_{guard_name}", help=f"מחק את {guard_name}"):
                st.session_state['deleted_guards'].add(guard_name)
                st.rerun()
        st.markdown("---")

    # Save selections back
    st.session_state['saved_name_selections'] = saved_selections

    matched_count = len(name_map)
    unmatched_count = len(active_matches) - matched_count
    c1, c2 = st.columns(2)
    c1.metric("שמות ששודכו", f"{matched_count} / {len(active_matches)}")
    c2.metric("🗑️ מאבטחים שנמחקו", len(deleted))

    if deleted:
        with st.expander("📋 מאבטחים שנמחקו (לחץ לשחזור)"):
            for d_name in sorted(deleted):
                rc1, rc2 = st.columns([5, 1])
                rc1.write(d_name)
                if rc2.button("♻️", key=f"restore_{d_name}", help="שחזר מאבטח"):
                    st.session_state['deleted_guards'].discard(d_name)
                    st.rerun()

    col_back, col_next = st.columns(2)
    with col_back:
        if st.button("→ חזור", use_container_width=True):
            # FIX #1: Save current name_map before going back
            st.session_state['name_map_final'] = name_map
            st.session_state['gap_step'] = 1
            st.rerun()
    with col_next:
        if st.button("המשך לשלב הבא ←", type="primary", use_container_width=True):
            st.session_state['name_map_final'] = name_map
            # Only auto-match positions if we haven't done it before
            if not st.session_state['pos_matches_multi']:
                pos_matches = auto_match_positions_multi(
                    st.session_state['positions_b'],
                    st.session_state['positions_a'],
                    master_pos_map=st.session_state.get('master_pos_map', {})
                )
                st.session_state['pos_matches_multi'] = pos_matches
            st.session_state['step_completed'].add(2)
            st.session_state['gap_step'] = 3
            st.rerun()

# ============================================================
# STEP 3: Position Matching (FIX #3: Multi-select)
# ============================================================
elif current_step == 3:
    st.header("🏢 שלב 3: שידוך עמדות / אתרים")
    st.info("שדך לכל עמדה ממאגר B **מספר** עמדות ממאגר A (בסיסי, מתקדם א, מתקדם ב, וכו.). ניתן גם למחוק עמדות מהרשימה.")

    positions_a = st.session_state['positions_a']
    positions_b = st.session_state['positions_b']
    pos_matches = st.session_state['pos_matches_multi']
    deleted_positions = st.session_state['deleted_positions']

    if not pos_matches and positions_b:
        pos_matches = auto_match_positions_multi(
            positions_b, 
            positions_a,
            master_pos_map=st.session_state.get('master_pos_map', {})
        )
        st.session_state['pos_matches_multi'] = pos_matches

    # Filter out deleted positions
    active_pos_matches = [m for m in pos_matches if m['source_b'] not in deleted_positions]

    pos_map_multi = {}  # {pos_b: [pos_a_1, pos_a_2, ...]}
    saved_pos_selections = st.session_state['saved_pos_selections']

    if active_pos_matches:
        for i, m in enumerate(active_pos_matches):
            pos_name = m['source_b']  # Use name as stable key (not index!)
            col_title, col_del = st.columns([9, 1])
            with col_title:
                st.markdown(f"### 🏢 {pos_name}")
                st.caption("עמדה ממאגר B (סידור עבודה)")
            with col_del:
                if st.button("🗑️", key=f"del_pos_{pos_name}", help=f"מחק את {pos_name}"):
                    st.session_state['deleted_positions'].add(pos_name)
                    st.rerun()

            # Multi-select for Source A positions
            # Use saved selection keyed by NAME (stable across deletions)
            if pos_name in saved_pos_selections:
                default_vals = [v for v in saved_pos_selections[pos_name] if v in positions_a]
            elif m['matched_a']:
                default_vals = [v for v in m['matched_a'] if v in positions_a]
            else:
                default_vals = []

            selected = st.multiselect(
                f"בחר עמדות מתאימות ממאגר A",
                positions_a,
                default=default_vals,
                key=f"pos_match_multi_{pos_name}",
                label_visibility="visible",
                help="ניתן לבחור מספר עמדות (בסיסי, מתקדם א, מתקדם ב וכו.)"
            )

            # Save selection keyed by name
            saved_pos_selections[pos_name] = selected

            if selected:
                pos_map_multi[pos_name] = selected
                st.success(f"✅ שודך ל-{len(selected)} עמדות: {', '.join(selected)}")

            conf = m['confidence']
            cls = 'conf-high' if conf >= 80 else ('conf-mid' if conf >= 50 else 'conf-low')
            st.markdown(f'ביטחון התאמה אוטומטית: <span class="{cls}">{conf}%</span>', unsafe_allow_html=True)
            st.markdown("---")

        st.session_state['saved_pos_selections'] = saved_pos_selections

        total_mapped = sum(len(v) for v in pos_map_multi.values())
        c1, c2 = st.columns(2)
        c1.metric("עמדות B ששודכו", f"{len(pos_map_multi)} / {len(active_pos_matches)}")
        c2.metric("🗑️ עמדות שנמחקו", len(deleted_positions))
    else:
        st.warning("לא נמצאו עמדות לשידוך.")

    # Restore deleted positions
    if deleted_positions:
        with st.expander("📋 עמדות שנמחקו (לחץ לשחזור)"):
            for d_pos in sorted(deleted_positions):
                rc1, rc2 = st.columns([5, 1])
                rc1.write(d_pos)
                if rc2.button("♻️", key=f"restore_pos_{d_pos}", help="שחזר עמדה"):
                    st.session_state['deleted_positions'].discard(d_pos)
                    st.rerun()

    col_back, col_next = st.columns(2)
    with col_back:
        if st.button("→ חזור", use_container_width=True):
            st.session_state['pos_map_multi'] = pos_map_multi
            st.session_state['gap_step'] = 2
            st.rerun()
    with col_next:
        if st.button("המשך לשלב הבא ←", type="primary", use_container_width=True):
            st.session_state['pos_map_multi'] = pos_map_multi
            st.session_state['step_completed'].add(3)
            st.session_state['gap_step'] = 4
            st.rerun()

# ============================================================
# STEP 4: Date Matching
# ============================================================
elif current_step == 4:
    st.header("📅 שלב 4: אימות תאריכים")

    dates_a = st.session_state['dates_a']
    dates_b = st.session_state['dates_b']

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("**תאריכים במאגר A (יל\"מ):**")
        for d in dates_a:
            st.write(d)
    with col2:
        st.markdown("**תאריכים במאגר B (סידור):**")
        for d in dates_b:
            st.write(d)

    common = set(dates_a) & set(dates_b)
    only_a = set(dates_a) - set(dates_b)
    only_b = set(dates_b) - set(dates_a)

    if common:
        st.success(f"✅ {len(common)} תאריכים משותפים")
    if only_a:
        st.warning(f"⚠️ {len(only_a)} תאריכים שמופיעים רק ביל\"מ: {', '.join(sorted(only_a))}")
    if only_b:
        st.warning(f"⚠️ {len(only_b)} תאריכים שמופיעים רק בסידור: {', '.join(sorted(only_b))}")
    if not common and (dates_a and dates_b):
        st.error("❌ אין תאריכים משותפים בין שני המאגרים!")

    col_back, col_next = st.columns(2)
    with col_back:
        if st.button("→ חזור", use_container_width=True):
            st.session_state['gap_step'] = 3
            st.rerun()
    with col_next:
        if st.button("🔍 הפעל זיהוי פערים ←", type="primary", use_container_width=True):
            st.session_state['step_completed'].add(4)
            st.session_state['gap_step'] = 5
            st.rerun()

# ============================================================
# STEP 5: Gap Detection & Results
# ============================================================
elif current_step == 5:
    st.header("🔍 שלב 5: תוצאות זיהוי פערים")

    df_a = st.session_state['df_a']
    df_b = st.session_state['df_b']
    name_map = st.session_state['name_map_final']
    pos_map_multi = st.session_state.get('pos_map_multi', {})
    deleted_positions = st.session_state.get('deleted_positions', set())

    if df_a is None or df_b is None:
        st.error("חסרים נתונים — חזור לשלב 1.")
    else:
        with st.spinner("מחפש פערים..."):
            gap_df = detect_gaps(df_a, df_b, name_map, pos_map_multi, deleted_positions)
            st.session_state['gap_results'] = gap_df

        if gap_df is not None and not gap_df.empty:
            # --- Summary Cards ---
            total = len(gap_df)
            ok_count = len(gap_df[gap_df['סטטוס'].str.contains('תקין')])
            partial_count = len(gap_df[gap_df['סטטוס'].str.contains('⚠️')])
            missing_count = len(gap_df[gap_df['סטטוס'].str.contains('❌')])

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("סה\"כ משמרות", total)
            c2.metric("✅ תקין", ok_count)
            c3.metric("⚠️ חלקי", partial_count)
            c4.metric("❌ חסר", missing_count)

            st.markdown("---")

            # --- Filter ---
            filter_option = st.radio(
                "סנן תוצאות:",
                ["הכל", "⚠️ חלקי בלבד", "❌ חסר בלבד", "בעיות בלבד (חלקי + חסר)"],
                horizontal=True
            )

            display_df = gap_df.copy()
            if filter_option == "⚠️ חלקי בלבד":
                display_df = display_df[display_df['סטטוס'].str.contains('⚠️')]
            elif filter_option == "❌ חסר בלבד":
                display_df = display_df[display_df['סטטוס'].str.contains('❌')]
            elif filter_option == "בעיות בלבד (חלקי + חסר)":
                display_df = display_df[~display_df['סטטוס'].str.contains('תקין')]

            # Show table without the internal 'עמדה ראשית' column
            show_cols = [c for c in display_df.columns if c != 'עמדה ראשית']
            st.dataframe(
                display_df[show_cols],
                use_container_width=True,
                height=500,
                column_config={
                    'סטטוס': st.column_config.TextColumn(width="medium"),
                    'תאריך': st.column_config.TextColumn(width="small"),
                    'שעות (סידור)': st.column_config.TextColumn(width="medium"),
                }
            )

            # --- Natural Language Report ---
            st.markdown("---")
            st.subheader("📝 דוח טקסטואלי (שפה טבעית)")

            natural_text = generate_natural_language_report(display_df)
            st.markdown(natural_text)

            # --- Rich Excel Export ---
            st.markdown("---")
            st.subheader("📥 ייצוא דוח")

            @st.cache_data(show_spinner=False)
            def get_cached_excel_bytes(df_obj, _text):
                buf = generate_rich_excel(df_obj, _text)
                return buf.getvalue()

            final_bytes = get_cached_excel_bytes(display_df, natural_text)

            st.download_button(
                label="📥 הורד דוח פערים מעוצב (Excel)",
                data=final_bytes,
                file_name="Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_btn_report"
            )
        else:
            st.warning("לא נמצאו תוצאות. ודא שהנתונים תקינים ושהשמות שודכו.")

    if st.button("→ חזור לשלב 4", use_container_width=True):
        st.session_state['gap_step'] = 4
        st.rerun()
