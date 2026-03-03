import streamlit as st
import pandas as pd
import re
import io
from datetime import datetime, timedelta
import holidays
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def load_css():
    css_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "style.css")
    if os.path.exists(css_path):
        with open(css_path, "r", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# --- 1. פונקציית ניקוי שם עמדה ---
def clean_location(original_loc):
    """מנקה את שם העמדה מתווים מיותרים ומחזיר את השם כפי שהוא מופיע בקובץ"""
    if not isinstance(original_loc, str) or not original_loc.strip():
        return ""
    return original_loc.replace('"', '').replace("'", "").strip()

# --- 3. לוגיקת שכר וחוקי עבודה ---

def get_israel_holidays(year):
    return holidays.IL(years=year)

def is_holiday(date_obj, il_holidays):
    """בודק אם התאריך הוא חג או ערב חג"""
    return date_obj in il_holidays

def is_weekend(date_obj, start_time_str):
    """
    בודק אם המשמרת נחשבת שבת:
    מיום שישי ב-16:00 עד יום ראשון ב-04:00
    """
    weekday = date_obj.weekday() # Monday=0, Sunday=6
    
    # יום שישי (4)
    if weekday == 4:
        if not start_time_str: return False
        try:
            h, m = map(int, start_time_str.split(':'))
            if h >= 16:
                return True
        except:
            return False
            
    # יום שבת (5)
    if weekday == 5:
        return True
        
    # יום ראשון (6) - עד 04:00 בבוקר לרוב נחשב משמרת לילה של שבת אם התחילה לפני
    # אבל לפי החוק היבש שבת נגמרת במוצ"ש. לצורך הפשטות והמקובל:
    # נבדוק אם זו משמרת שהתחילה בשבת או בשישי בערב.
    # אם המשמרת *נופלת* על השעות האלו.
    return False

def calculate_hours(start_str, end_str):
    """מחשב שעות עבודה בין שתי שעות"""
    if not start_str or not end_str:
        return 0.0
    
    try:
        fmt = '%H:%M'
        t1 = datetime.strptime(start_str, fmt)
        t2 = datetime.strptime(end_str, fmt)
        
        if t2 < t1:
            t2 += timedelta(days=1)
            
        diff = t2 - t1
        return diff.total_seconds() / 3600.0
    except:
        return 0.0

def calculate_shift_pay(shift, wage, il_holidays):
    """
    מבצע חישוב שכר מפורט למשמרת בודדת
    """
    start_str = shift['Start']
    end_str = shift['End']
    date_obj = shift['dt'].date()
    
    total_hours = calculate_hours(start_str, end_str)
    if total_hours <= 0:
        return {
            "hours_100": 0, "hours_125": 0, "hours_150": 0, 
            "hours_175": 0, "hours_200": 0, "total_pay": 0,
            "category": "שגיאה"
        }

    # זיהוי סוג משמרת
    is_hol = is_holiday(date_obj, il_holidays)
    is_wknd = is_weekend(date_obj, start_str)

    # קביעת סף שעות נוספות
    regular_threshold = 8.0
         
    if is_wknd or is_hol: # שבת/חג: אין "שעות רגילות" במובן של 100%, הכל 150% ומעלה
        regular_threshold = 0 # הכל נחשב מיוחד
        
    # --- חישוב השעות לפי מדרגות ---
    # מדרגות:
    # רגיל: 100% עד הסף
    # שעות נוספות (פשוטות): שנתיים ראשונות 125%, מעבר לכך 150%
    # שבת/חג: הכל 150% (בסיס). נוספות: שעתיים ראשונות 175%, מעבר 200%
    
    h_100 = 0
    h_125 = 0
    h_150 = 0
    h_175 = 0
    h_200 = 0
    
    rem_hours = total_hours
    
    if is_wknd or is_hol:
        # בסיס שבת/חג הוא 150%
        # אבל יש מגבלה יומית לשעות ה"רגילות" של שבת?
        # לפי החוק: שעות נוספות בשבת מחושבות מהשעה הראשונה? לא, מהשעה ה-8/9/7.
        # אז עד הסף (למשל 8 שעות) זה 150%. 
        # שעה 9-10 זה 175%.
        # שעה 11+ זה 200%.
        # סף לשבת: נקבע לפי יום רגיל (8/9/7)
        # אם שבת בערב מחשיבים 7? לרוב כן. נלך על מחמיר: 8 שעות או 7 אם לילה.
        
        limit_base = 8.0
        
        # תיקון לחוק: בשבת הכל 150%. שעות נוספות זה על ה-150%.
        # אז עד limit_base -> 150%
        # limit_base עד limit_base+2 -> 175%
        # מעבר -> 200%
        
        base_sat = min(rem_hours, limit_base)
        h_150 += base_sat
        rem_hours -= base_sat
        
        ot_1 = min(rem_hours, 2.0)
        h_175 += ot_1
        rem_hours -= ot_1
        
        h_200 += rem_hours
        
        category = "שבת/חג"
        
    else: # יום חול
        # עד הסף -> 100%
        base_reg = min(rem_hours, regular_threshold)
        h_100 += base_reg
        rem_hours -= base_reg
        
        # שעות נוספות
        ot_1 = min(rem_hours, 2.0)
        h_125 += ot_1
        rem_hours -= ot_1
        
        h_150 += rem_hours # כל השאר 150%
        
        category = "רגיל"

    # חישוב כספי
    pay = (h_100 * wage * 1.00) + \
          (h_125 * wage * 1.25) + \
          (h_150 * wage * 1.50) + \
          (h_175 * wage * 1.75) + \
          (h_200 * wage * 2.00)
          
    return {
        "hours_100": h_100,
        "hours_125": h_125,
        "hours_150": h_150,
        "hours_175": h_175,
        "hours_200": h_200,
        "total_pay": pay,
        "category": category
    }

# --- 4. פונקציית פענוח הקבצים ---
def parse_schedule_file(uploaded_file, employee_name):
    try:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, header=None)
        else:
            df = pd.read_excel(uploaded_file, header=None)
    except Exception as e:
        st.error(f"שגיאה בטעינת הקובץ {uploaded_file.name}: {e}")
        return []

    shifts = []
    
    date_row_idx = -1
    dates_map = {} 
    
    for i, row in df.iterrows():
        row_str = " ".join([str(x) for x in row if pd.notna(x)])
        if re.search(r'\d{1,2}[/.]\d{1,2}[/.]\d{4}', row_str):
            date_row_idx = i
            for c in range(len(df.columns)):
                val = str(df.iloc[i, c])
                match = re.search(r'(\d{1,2}[/.]\d{1,2}[/.]\d{4})', val)
                if match:
                    date_str = match.group(1)
                    dates_map[c] = date_str
                    if c+1 < len(df.columns):
                        dates_map[c+1] = date_str
            break
            
    if date_row_idx == -1:
        return []

    current_location = ""
    IGNORE_LOCS = ["בוקר", "צהריים", "לילה", "מתגברים", "רענון", "nan", "None", ""]

    for i in range(date_row_idx + 1, len(df)):
        row = df.iloc[i]
        
        col_a_val = str(row[0]).strip()
        if col_a_val and col_a_val not in IGNORE_LOCS and not re.search(r'\d{1,2}:\d{2}', col_a_val):
            current_location = col_a_val
            
        for c in range(1, len(df.columns)):
            cell_val = str(row[c])
            clean_cell_val = re.sub(r'\s*\(.*?\)', '', cell_val).replace('*', '').strip()
            
            if employee_name in cell_val or employee_name in clean_cell_val:
                if c not in dates_map:
                    continue 
                
                date_found = dates_map[c]
                start_time = ""
                end_time = ""
                
                if i + 1 < len(df):
                    next_row = df.iloc[i+1]
                    t1 = str(next_row[c]).strip()
                    t2 = str(next_row[c+1] if c+1 < len(df.columns) else "").strip()
                    if re.match(r'\d{1,2}:\d{2}', t1) or re.match(r'\d{1,2}:\d{2}', t2):
                        end_time = t1
                        start_time = t2 
                        
                if not start_time and i - 1 >= 0:
                    prev_row = df.iloc[i-1]
                    t1 = str(prev_row[c]).strip()
                    t2 = str(prev_row[c+1] if c+1 < len(df.columns) else "").strip()
                    if re.match(r'\d{1,2}:\d{2}', t1) or re.match(r'\d{1,2}:\d{2}', t2):
                        end_time = t1
                        start_time = t2

                final_loc = current_location
                if not final_loc:
                    for k in range(i, date_row_idx, -1):
                        val_up = str(df.iloc[k, 0]).strip()
                        if val_up and val_up not in IGNORE_LOCS:
                            final_loc = val_up
                            break
                            
                shifts.append({
                    "Date": date_found,
                    "Raw_Loc": final_loc,
                    "Start": start_time,
                    "End": end_time,
                    "Source_File": uploaded_file.name
                })

    return shifts

def extract_employee_ids(uploaded_file):
    """מחלץ מילון של שם עובד -> ת.ז מתוך קובץ"""
    try:
        uploaded_file.seek(0)
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
            
        name_col = None
        id_col = None
        for col in df.columns:
            c = str(col).lower()
            if "שם" in c or "name" in c:
                name_col = col
            if "ת.ז" in c or "תז" in c or "זהות" in c or "id" in c:
                id_col = col
                
        if name_col and id_col:
            mapping = {}
            for _, row in df.iterrows():
                name = str(row[name_col]).strip()
                id_val = str(row[id_col]).strip()
                if name and str(name) != "nan" and id_val and str(id_val) != "nan":
                    clean_name = re.sub(r'\s*\(.*?\)', '', name).replace('*', '').strip()
                    mapping[clean_name] = id_val
            return mapping
    except:
        pass
    return {}

def extract_all_locations(uploaded_file):
    """מחלץ את כל שמות העמדות מקובץ סידור, ללא תלות בעובד מסוים"""
    try:
        uploaded_file.seek(0)
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, header=None)
        else:
            df = pd.read_excel(uploaded_file, header=None)
    except:
        return set()

    locations = set()
    IGNORE_LOCS = ["בוקר", "צהריים", "לילה", "מתגברים", "רענון", "nan", "None", ""]

    # מוצא את שורת התאריכים
    date_row_idx = -1
    for i, row in df.iterrows():
        row_str = " ".join([str(x) for x in row if pd.notna(x)])
        if re.search(r'\d{1,2}[/.]\d{1,2}[/.]\d{4}', row_str):
            date_row_idx = i
            break

    if date_row_idx == -1:
        return set()

    # סורק את עמודה A אחרי שורת התאריכים ומוצא שמות עמדות
    for i in range(date_row_idx + 1, len(df)):
        col_a_val = str(df.iloc[i, 0]).strip()
        if col_a_val and col_a_val not in IGNORE_LOCS and not re.search(r'\d{1,2}:\d{2}', col_a_val):
            cleaned = clean_location(col_a_val)
            if cleaned:
                locations.add(cleaned)

    return locations

def extract_all_employees(uploaded_file):
    """מחלץ את כל שמות העובדים מקובץ סידור"""
    try:
        uploaded_file.seek(0)
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file, header=None)
        else:
            df = pd.read_excel(uploaded_file, header=None)
    except:
        return set()

    employees = set()
    IGNORE_LOCS = ["בוקר", "צהריים", "לילה", "מתגברים", "רענון", "nan", "None", ""]

    # מוצא את שורת התאריכים
    date_row_idx = -1
    for i, row in df.iterrows():
        row_str = " ".join([str(x) for x in row if pd.notna(x)])
        if re.search(r'\d{1,2}[/.]\d{1,2}[/.]\d{4}', row_str):
            date_row_idx = i
            break

    if date_row_idx == -1:
        return set()

    # סורק את כל התאים אחרי שורת התאריכים ומוצא שמות עובדים
    for i in range(date_row_idx + 1, len(df)):
        col_a_val = str(df.iloc[i, 0]).strip()
        # אם זו שורת עמדה/כותרת - נדלג
        is_location_row = (col_a_val and col_a_val not in IGNORE_LOCS 
                          and not re.search(r'\d{1,2}:\d{2}', col_a_val))
        
        for c in range(1, len(df.columns)):
            cell_val = str(df.iloc[i, c]).strip()
            # מסנן ערכים שלא יכולים להיות שמות עובדים
            if (cell_val and cell_val != "nan" and cell_val != "None"
                and not re.match(r'^\d{1,2}:\d{2}$', cell_val)  # לא שעה
                and not re.match(r'^\d+\.?\d*$', cell_val)  # לא מספר
                and not re.search(r'\d{1,2}[/.]\d{1,2}[/.]\d{4}', cell_val)  # לא תאריך
                and len(cell_val) > 1  # לא תו בודד
                and cell_val not in IGNORE_LOCS
                and cell_val not in ["בוקר", "צהריים", "לילה"]
            ):
                # ניקוי סוגריים ותוכנם מהשם, וגם מחיקת כוכביות
                clean_name = re.sub(r'\s*\(.*?\)', '', cell_val).replace('*', '').strip()
                if clean_name:
                    employees.add(clean_name)

    return employees

def generate_report_for_employee(employee_name, files, month, year, hourly_wage, travel_rate):
    """מייצר דוח שעות ושכר עבור עובד ספציפי"""
    all_shifts = []
    for file in files:
        file.seek(0)
        shifts = parse_schedule_file(file, employee_name)
        all_shifts.extend(shifts)

    processed_shifts = []
    seen_keys = set()

    for s in all_shifts:
        try:
            clean_date = s['Date'].replace('.', '/')
            dt_obj = pd.to_datetime(clean_date, dayfirst=True)

            if dt_obj.month == month and dt_obj.year == year:
                unique_key = f"{dt_obj}_{s['Start']}_{s['End']}_{s['Raw_Loc']}"

                if unique_key not in seen_keys:
                    s['dt'] = dt_obj
                    processed_shifts.append(s)
                    seen_keys.add(unique_key)
        except:
            continue

    # הסרת כפילויות "מתגבר" - אם יש משמרת באותו תאריך ושעות
    # גם תחת "מתגבר"/"מתגברים" וגם תחת עמדה ספציפית, נשמור רק את העמדה הספציפית
    REINFORCEMENT_NAMES = ["מתגבר", "מתגברים"]
    
    shifts_to_remove = set()
    for i, s in enumerate(processed_shifts):
        loc = clean_location(s['Raw_Loc'])
        if loc in REINFORCEMENT_NAMES:
            # בודקים אם יש משמרת אחרת באותו תאריך ושעות עם עמדה ספציפית
            for j, other in enumerate(processed_shifts):
                if i != j and j not in shifts_to_remove:
                    other_loc = clean_location(other['Raw_Loc'])
                    if (s['dt'].date() == other['dt'].date()
                        and s['Start'] == other['Start']
                        and s['End'] == other['End']
                        and other_loc not in REINFORCEMENT_NAMES):
                        # נמצאה עמדה ספציפית - נסמן את ה"מתגבר" למחיקה
                        shifts_to_remove.add(i)
                        break
    
    processed_shifts = [s for i, s in enumerate(processed_shifts) if i not in shifts_to_remove]

    start_date = f"{year}-{month}-01"
    end_date = pd.Period(f"{year}-{month}").end_time.date()
    date_range = pd.date_range(start_date, end_date)

    il_holidays = get_israel_holidays(year)

    final_rows = []
    days_heb = {6: "ראשון", 0: "שני", 1: "שלישי", 2: "רביעי", 3: "חמישי", 4: "שישי", 5: "שבת"}
    days_num = {6: 1, 0: 2, 1: 3, 2: 4, 3: 5, 4: 6, 5: 7}

    for d in date_range:
        shifts_for_day = [s for s in processed_shifts if s['dt'].date() == d.date()]

        d_display = d.strftime('%d/%m/%Y')
        wd = d.weekday()

        if not shifts_for_day:
            final_rows.append({
                "תאריך": d_display,
                "יום בשבוע": days_num[wd],
                "יום שם": days_heb[wd],
                "עמדת תגבור": "",
                "תעריף נסיעות יומי": 0.0,
                "שעת כניסה": "",
                "שעת יציאה": "",
                "סהמ שעות": 0,
                "שעות 100%": 0,
                "שעות 125%": 0,
                "שעות 150%": 0,
                "שעות 175%": 0,
                "שעות 200%": 0,
                "שכר יומי": 0,
                "סוג יום": ""
            })
        else:
            shifts_for_day.sort(key=lambda x: x['Start'] if x['Start'] else "00:00")

            for shift in shifts_for_day:
                loc = clean_location(shift['Raw_Loc'])

                travel_val = travel_rate if travel_rate > 0 else 0.0

                final_rows.append({
                    "תאריך": d_display,
                    "יום בשבוע": days_num[wd],
                    "יום שם": days_heb[wd],
                    "עמדת תגבור": loc,
                    "תעריף נסיעות יומי": travel_val,
                    "שעת כניסה": shift['Start'],
                    "שעת יציאה": shift['End'],
                    "סהמ שעות": 0, "שעות 100%": 0, "שעות 125%": 0,
                    "שעות 150%": 0, "שעות 175%": 0, "שעות 200%": 0,
                    "שכר יומי": 0, "סוג יום": ""
                })

                row_data = final_rows[-1]

                pay_data = calculate_shift_pay(shift, hourly_wage, il_holidays)

                row_data.update({
                    "סהמ שעות": round(pay_data["hours_100"] + pay_data["hours_125"] + pay_data["hours_150"] + pay_data["hours_175"] + pay_data["hours_200"], 2),
                    "שעות 100%": round(pay_data["hours_100"], 2),
                    "שעות 125%": round(pay_data["hours_125"], 2),
                    "שעות 150%": round(pay_data["hours_150"], 2),
                    "שעות 175%": round(pay_data["hours_175"], 2),
                    "שעות 200%": round(pay_data["hours_200"], 2),
                    "שכר יומי": round(pay_data["total_pay"] + (travel_rate if travel_rate else 0), 2),
                    "סוג יום": pay_data["category"]
                })

    df_final = pd.DataFrame(final_rows)
    total_salary = df_final["שכר יומי"].sum()

    return df_final, processed_shifts, total_salary

# --- 5. הממשק הראשי ---
def main():
    st.set_page_config(page_title="מערכת סידור עבודה חכם", layout="wide", page_icon="📅")
    load_css()
    
    st.title("👮‍♂️ מחולל דוחות שעות - אבטחה")
    
    with st.sidebar:
        st.header("הגדרות חיפוש")
        files = st.file_uploader("העלה קבצי סידור (CSV/Excel)", accept_multiple_files=True, type=['csv', 'xlsx'])
        
        id_file = st.file_uploader("העלה קובץ מספרי תעודת זהות (אופציונלי)", type=['csv', 'xlsx'])
        
        st.markdown("---")
        st.subheader("תאריך הדוח")
        
        current_year = 2026
        current_month = 1
        col1, col2 = st.columns(2)
        with col1:
            month = st.number_input("חודש", min_value=1, max_value=12, value=current_month)
        with col2:
            year = st.number_input("שנה", min_value=2020, max_value=2030, value=current_year)

    # --- שלב 1: חילוץ שמות עובדים ותעודות זהות מהקבצים ---
    employee_ids_map = {}
    if id_file:
        employee_ids_map = extract_employee_ids(id_file)
        if employee_ids_map:
            st.sidebar.success(f"נטענו {len(employee_ids_map)} מספרי ת.ז")
        else:
            st.sidebar.warning("לא זוהו עמודות שם ות.ז בקובץ זה")

    if files:
        # חילוץ כל העמדות וכל העובדים
        all_locations = set()
        all_employees = set()
        
        for file in files:
            file.seek(0)
            all_locations.update(extract_all_locations(file))
            file.seek(0)
            all_employees.update(extract_all_employees(file))
        
        all_employees_sorted = sorted(all_employees)
        all_locations_sorted = sorted(all_locations)
        
        # הצגת רשימת עמדות
        with st.expander(f"📍 כל העמדות הקיימות ({len(all_locations_sorted)})", expanded=False):
            for idx, loc in enumerate(all_locations_sorted, 1):
                st.write(f"{idx}. {loc}")
        
        st.markdown("---")
        
        if "wage_groups" not in st.session_state:
            st.session_state.wage_groups = {
                "ברירת מחדל": {"wage": 32.0, "travel": 0.0, "employees": []}
            }

        # --- שלב 2: יצירת ושיוך לקבוצות תנאים ---
        st.subheader("💼 ניהול קבוצות שכר ושיוך עובדים")
        st.caption("יצירת קבוצות תנאים ושיוך עובדים לכל סט (שכר שעתי, החזר נסיעות)")

        # הוספת קבוצה חדשה
        with st.expander("➕ הוספת קבוצת שכר חדשה", expanded=False):
            col_g1, col_g2, col_g3 = st.columns([2, 1, 1])
            new_g_name = col_g1.text_input("שם הקבוצה (למשל: אחמ\"שים)")
            new_g_wage = col_g2.number_input("שכר שעתי", min_value=29.0, value=32.0, step=0.5, key="new_g_wage")
            new_g_travel = col_g3.number_input("נסיעות יומי", min_value=0.0, value=0.0, step=0.5, key="new_g_travel")
            if st.button("הוסף קבוצה"):
                if new_g_name and new_g_name not in st.session_state.wage_groups:
                    st.session_state.wage_groups[new_g_name] = {"wage": new_g_wage, "travel": new_g_travel, "employees": []}
                    st.success(f"קבוצה '{new_g_name}' נוספה!")
                    st.rerun()
                elif new_g_name in st.session_state.wage_groups:
                    st.warning("שם קבוצה זה כבר קיים.")
        
        st.markdown("---")
        
        # UI לכל קבוצה קיימת
        for g_name in list(st.session_state.wage_groups.keys()):
            with st.container():
                st.markdown(f"**קבוצה: {g_name}**")
                col_w, col_t, col_del = st.columns([1.5, 1.5, 1])
                new_wage = col_w.number_input("שכר שעתי", min_value=0.0, value=float(st.session_state.wage_groups[g_name]["wage"]), step=0.5, key=f"w_{g_name}")
                new_travel = col_t.number_input("נסיעות יומי", min_value=0.0, value=float(st.session_state.wage_groups[g_name]["travel"]), step=0.5, key=f"t_{g_name}")
                
                 # עדכון ערכים בזמן אמת
                st.session_state.wage_groups[g_name]["wage"] = new_wage
                st.session_state.wage_groups[g_name]["travel"] = new_travel
                
                if g_name != "ברירת מחדל":
                     # יישור הכפתור מחיקה לשורה למטה
                     st.write("") # מרווח קטן
                     if col_del.button("🗑️ מחק קבוצה", key=f"del_{g_name}"):
                         del st.session_state.wage_groups[g_name]
                         st.rerun()
                
                # בחירת עובדים לקבוצה
                options = sorted(list(set(all_employees_sorted) | set(st.session_state.wage_groups[g_name]["employees"])))
                selected_emps = st.multiselect(
                    "שייך עובדים:", 
                    options=options, 
                    default=[e for e in st.session_state.wage_groups[g_name]["employees"] if e in options], 
                    key=f"emps_{g_name}",
                    placeholder="בחר עובדים מתוך הרשימה..."
                )
                st.session_state.wage_groups[g_name]["employees"] = selected_emps
                st.markdown("---")
        
        # איסוף כל העובדים שנבחרו על מנת להפיק דוחות
        employees_to_process = []
        emp_to_group = {}
        for g_name, g_data in st.session_state.wage_groups.items():
            for emp in g_data["employees"]:
                # אם עובד נבחר בכמה קבוצות, הקבוצה האחרונה תדרוס את קודמתה
                if emp not in employees_to_process:
                    employees_to_process.append(emp)
                emp_to_group[emp] = {"wage": g_data["wage"], "travel": g_data["travel"]}
                
        if employees_to_process:
            st.info(f"סה\"כ {len(employees_to_process)} עובדים שויכו ומוכנים להפקת דוח.")
            process_btn = st.button("🚀 הפק דוחות לעובדים ששויכו", type="primary")
        else:
            st.warning("שייך לפחות עובד אחד לאחת מהקבוצות כדי להפיק דוחות.")
            process_btn = False
        
        # --- שלב 3: הפקת דוחות ---
        if process_btn and employees_to_process:
            progress_bar = st.progress(0)
            
            all_excel_sheets = {}
            
            for emp_idx, emp_name in enumerate(employees_to_process):
                g_data = emp_to_group[emp_name]
                df_final, processed_shifts, total_salary = generate_report_for_employee(
                    emp_name, files, month, year, g_data["wage"], g_data["travel"]
                )
                
                progress_bar.progress((emp_idx + 1) / len(employees_to_process))
                
                # דוח לכל עובד
                st.markdown(f"## 📋 דוח עבור: **{emp_name}**")
                
                if len(processed_shifts) == 0:
                    st.warning(f"לא נמצאו משמרות עבור {emp_name} בחודש {month}/{year}")
                else:
                    st.success(f"נמצאו {len(processed_shifts)} משמרות עבור {emp_name} בחודש {month}/{year}")
                    
                    # עמדות שבהן העובד עבד
                    employee_locations = sorted(set(
                        clean_location(s['Raw_Loc']) for s in processed_shifts if clean_location(s['Raw_Loc'])
                    ))
                    
                    with st.expander(f"📌 עמדות של {emp_name} ({len(employee_locations)})", expanded=False):
                        for idx, loc in enumerate(employee_locations, 1):
                            st.write(f"{idx}. {loc}")
                    
                    # הוספת שורת סיכום לתוך הטבלה
                    summary_row = {
                        "תאריך": "סה\"כ",
                        "יום בשבוע": None,
                        "יום שם": "",
                        "עמדת תגבור": emp_name,
                        "תעריף נסיעות יומי": round(df_final["תעריף נסיעות יומי"].sum(), 2),
                        "שעת כניסה": "",
                        "שעת יציאה": "",
                        "סהמ שעות": round(df_final["סהמ שעות"].sum(), 2),
                        "שעות 100%": round(df_final["שעות 100%"].sum(), 2),
                        "שעות 125%": round(df_final["שעות 125%"].sum(), 2),
                        "שעות 150%": round(df_final["שעות 150%"].sum(), 2),
                        "שעות 175%": round(df_final["שעות 175%"].sum(), 2),
                        "שעות 200%": round(df_final["שעות 200%"].sum(), 2),
                        "שכר יומי": round(total_salary, 2),
                        "סוג יום": "💰 סה\"כ לתשלום"
                    }
                    
                    df_display = df_final.copy()
                    
                    # הוספת ת.ז אם קיים
                    emp_id = employee_ids_map.get(emp_name, "")
                    if emp_id:
                        id_row = {col: "" for col in df_final.columns}
                        id_row["תאריך"] = "תעודת זהות:"
                        id_row["יום בשבוע"] = None
                        id_row["יום שם"] = emp_id
                        df_display = pd.concat([pd.DataFrame([id_row]), df_display], ignore_index=True)
                    
                    df_display = pd.concat([df_display, pd.DataFrame([summary_row])], ignore_index=True)
                    
                    st.dataframe(df_display, use_container_width=True)
                
                # שמירה לאקסל
                all_excel_sheets[emp_name] = df_display
                
                st.markdown("---")
            
            # כפתור הורדת כל הדוחות באקסל אחד
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                for emp_name, df in all_excel_sheets.items():
                    # שם הגיליון מוגבל ל-31 תווים באקסל, ואסור שיכיל תווים מיוחדים
                    sheet_name = re.sub(r'[\\/*?:\[\]]', '', emp_name)[:31]
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
                    
                    # --- Excel Styling ---
                    worksheet = writer.sheets[sheet_name]
                    worksheet.sheet_view.rightToLeft = True
                    
                    # Define Styles
                    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    summary_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    summary_font = Font(bold=True)
                    id_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    id_font = Font(bold=True, color="375623")
                    
                    thin_border = Border(
                        left=Side(style='thin', color='BFBFBF'),
                        right=Side(style='thin', color='BFBFBF'),
                        top=Side(style='thin', color='BFBFBF'),
                        bottom=Side(style='thin', color='BFBFBF')
                    )
                    center_align = Alignment(horizontal="center", vertical="center")
                    
                    # Style Header Row
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                        cell.border = thin_border
                        
                    # Find total rows and max columns
                    max_row = worksheet.max_row
                    max_col = worksheet.max_column
                    
                    # Format Data cells
                    for row_idx in range(2, max_row + 1):
                        is_id_row = False
                        is_summary_row = False
                        
                        header_val = str(worksheet.cell(row=row_idx, column=1).value or "")
                        if header_val == "תעודת זהות:":
                            is_id_row = True
                        elif header_val == "סה\"כ":
                            is_summary_row = True
                            
                        for col_idx in range(1, max_col + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = center_align
                            
                            if is_id_row:
                                cell.fill = id_fill
                                cell.font = id_font
                            elif is_summary_row:
                                cell.fill = summary_fill
                                cell.font = summary_font
                                
                    # Auto-adjust column width
                    for col_idx in range(1, max_col + 1):
                        max_length = 0
                        col_letter = get_column_letter(col_idx)
                        for cell in worksheet[col_letter]:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        worksheet.column_dimensions[col_letter].width = adjusted_width
            
            st.download_button(
                label="📥 הורד את כל הדוחות (Excel)",
                data=output.getvalue(),
                file_name=f"shift_reports_{month}_{year}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
