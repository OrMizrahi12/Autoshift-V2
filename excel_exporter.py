import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_styled_excel(roster, sorted_days, unique_positions):
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True # RTL support for Hebrew
    ws.title = "סידור עבודה"
    
    # Define styles
    header_fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    
    day_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    day_font = Font(bold=True, color="344054")
    
    shift_badge_fills = {
        "בוקר": PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid"),
        "צהריים": PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid"),
        "לילה": PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    }
    shift_font = Font(bold=True, color="FFFFFF")
    
    regular_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    shortage_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
    
    worker_font = Font(color="1D2939", bold=True)
    shortage_font = Font(color="C2410C", bold=True)
    
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color="D0D5DD"),
        right=Side(style='thin', color="D0D5DD"),
        top=Side(style='thin', color="D0D5DD"),
        bottom=Side(style='thin', color="D0D5DD")
    )
    
    shift_groups = [
        ("בוקר", ['M', 'DM']),
        ("צהריים", ['A']),
        ("לילה", ['N', 'DN']),
    ]
    
    def get_time_range(raw_s):
        if raw_s == 'M': return "07:00 - 15:00"
        if raw_s == 'A': return "15:00 - 23:00"
        if raw_s == 'N': return "23:00 - 07:00"
        if raw_s == 'DM': return "07:00 - 19:00"
        if raw_s == 'DN': return "19:00 - 07:00"
        if raw_s == 'SHORTAGE': return ""
        return ""

    current_row = 1
    
    for pos in unique_positions:
        pos_df = roster[roster['עמדה'] == pos]
        
        # Row: Position Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sorted_days) + 1)
        cell = ws.cell(row=current_row, column=1, value=f"🛡️ {pos}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        for col_idx in range(1, len(sorted_days) + 2):
            ws.cell(row=current_row, column=col_idx).border = thin_border
        current_row += 1
        
        # Row: Day Headers
        ws.cell(row=current_row, column=1, value="משמרת").fill = day_fill
        ws.cell(row=current_row, column=1).font = day_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        
        for col_idx, d in enumerate(sorted_days, start=2):
            c = ws.cell(row=current_row, column=col_idx, value=d)
            c.fill = day_fill
            c.font = day_font
            c.alignment = center_align
            c.border = thin_border
        current_row += 1
        
        # Rows: Shift Groups
        for group_label, raw_codes in shift_groups:
            day_workers = {}
            max_depth = 0
            
            for d in sorted_days:
                workers_regular = pos_df[
                    (pos_df['יום'] == d) &
                    (pos_df['raw_shift'].isin(raw_codes))
                ]
                shortage_rows = pos_df[
                    (pos_df['יום'] == d) &
                    (pos_df['raw_shift'] == 'SHORTAGE')
                ]
                relevant_shortages = []
                for _, sr in shortage_rows.iterrows():
                    shift_desc = str(sr['משמרת']).lower()
                    if group_label == "בוקר" and 'בוקר' in shift_desc:
                        relevant_shortages.append(sr)
                    elif group_label == "צהריים" and 'צהריים' in shift_desc:
                        relevant_shortages.append(sr)
                    elif group_label == "לילה" and 'לילה' in shift_desc:
                        relevant_shortages.append(sr)
                
                entries = []
                for _, w in workers_regular.iterrows():
                    time_str = get_time_range(w['raw_shift'])
                    entries.append({
                        'name': w['עובד'],
                        'time': time_str,
                        'is_shortage': False
                    })
                for sr in relevant_shortages:
                    entries.append({
                        'name': sr['עובד'],
                        'time': sr['משמרת'],
                        'is_shortage': True
                    })
                
                day_workers[d] = entries
                max_depth = max(max_depth, len(entries))
            
            if max_depth == 0:
                continue
                
            for i in range(max_depth):
                # Shift label cell
                if i == 0:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + max_depth - 1, end_column=1)
                    c = ws.cell(row=current_row, column=1, value=group_label)
                    c.fill = shift_badge_fills.get(group_label, day_fill)
                    c.font = shift_font
                    c.alignment = center_align
                
                # We need to manually add borders for all cells in the merged range
                for mr in range(current_row, current_row + max_depth):
                    ws.cell(row=mr, column=1).border = thin_border
                
                # Day cells
                for col_idx, d in enumerate(sorted_days, start=2):
                    entries = day_workers.get(d, [])
                    c = ws.cell(row=current_row, column=col_idx)
                    c.alignment = center_align
                    c.border = thin_border
                    
                    if i < len(entries):
                        entry = entries[i]
                        c.value = f"{entry['name']}\n{entry['time']}"
                        if entry['is_shortage']:
                            c.fill = shortage_fill
                            c.font = shortage_font
                        else:
                            c.fill = regular_fill
                            c.font = worker_font
                    else:
                        c.fill = regular_fill
                
                current_row += 1
                
        current_row += 1 # Empty row between positions

    # Auto-adjust column widths
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i) # Get the column name safely
        for cell in col:
            try:
                # Add a bit more padding for Hebrew names and times
                if len(str(cell.value).split('\n')[0]) > max_length:
                    max_length = len(str(cell.value).split('\n')[0])
            except:
                pass
        adjusted_width = (max_length + 6)
        if adjusted_width < 14: adjusted_width = 14
        if column == 'A': adjusted_width = 16
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
